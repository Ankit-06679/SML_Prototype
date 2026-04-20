"""
Exports analyzed_posts.json to a presentation-ready Excel workbook.

Sheets:
  1. Summary          — KPIs and overall breakdown
  2. All Posts        — every post with key fields
  3. Patients         — filtered by stakeholder
  4. Physicians
  5. Pharmacists
  6. Caregivers
  7. Payers
  8. Patient Advocacy Groups
"""
import json
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from collections import Counter
from datetime import datetime

INPUT_PATH  = "data/analyzed_posts.json"
OUTPUT_PATH = "data/SML_Mepolizumab_Report.xlsx"

STAKEHOLDERS = [
    "Patients", "Physicians", "Pharmacists",
    "Caregivers", "Payers", "Patient Advocacy Groups",
]

SENTIMENT_COLORS = {
    "Positive": "C6EFCE",
    "Negative": "FFC7CE",
    "Neutral":  "FFEB9C",
}

HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT  = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT   = Font(color="1F4E79", bold=True, size=13)
BORDER_SIDE  = Side(style="thin", color="BFBFBF")
CELL_BORDER  = Border(
    left=BORDER_SIDE, right=BORDER_SIDE,
    top=BORDER_SIDE,  bottom=BORDER_SIDE
)


def load_data():
    with open(INPUT_PATH, "r", encoding="utf-8") as f:
        return json.load(f)


def flatten_posts(posts: list) -> pd.DataFrame:
    rows = []
    for p in posts:
        rows.append({
            "Post ID":        p.get("id", ""),
            "Subreddit":      p.get("subreddit", ""),
            "Stakeholder":    p.get("stakeholder", ""),
            "Sentiment":      p.get("sentiment", ""),
            "Sentiment Score":round(float(p.get("sentiment_score", 0)), 2),
            "Emotion":        p.get("emotion", ""),
            "Themes":         ", ".join(p.get("themes", [])),
            "Drugs Mentioned":", ".join(p.get("drugs_mentioned", [])),
            "Key Entities":   ", ".join(p.get("key_entities", [])),
            "Summary":        p.get("summary", ""),
            "Representative Quote": p.get("quote", ""),
            "Post Score":     p.get("score", 0),
            "Num Comments":   p.get("num_comments", 0),
            "Date":           p.get("created_utc", "")[:10],
            "URL":            p.get("url", ""),
        })
    return pd.DataFrame(rows)


def style_header_row(ws, row: int, num_cols: int):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = CELL_BORDER


def style_data_rows(ws, start_row: int, end_row: int, num_cols: int, df: pd.DataFrame):
    sentiment_col = None
    for col in range(1, num_cols + 1):
        if ws.cell(row=start_row - 1, column=col).value == "Sentiment":
            sentiment_col = col
            break

    for row in range(start_row, end_row + 1):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = CELL_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            # Colour sentiment column
            if sentiment_col and col == sentiment_col:
                val = cell.value or ""
                color = SENTIMENT_COLORS.get(val, "FFFFFF")
                cell.fill = PatternFill("solid", fgColor=color)
        # Alternate row shading
        if row % 2 == 0:
            for col in range(1, num_cols + 1):
                c = ws.cell(row=row, column=col)
                if not c.fill or c.fill.fgColor.rgb in ("00000000", "FFFFFFFF"):
                    c.fill = PatternFill("solid", fgColor="F2F2F2")


def auto_col_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value or "")
                max_len = max(max_len, min(len(val), 60))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max(12, max_len + 2)


def write_df_to_sheet(ws, df: pd.DataFrame, title: str = ""):
    if title:
        ws.cell(row=1, column=1, value=title).font = TITLE_FONT
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="left")
        header_row = 2
    else:
        header_row = 1

    # Write headers
    for col_idx, col_name in enumerate(df.columns, start=1):
        ws.cell(row=header_row, column=col_idx, value=col_name)
    style_header_row(ws, header_row, len(df.columns))

    # Write data
    for row_idx, row in enumerate(df.itertuples(index=False), start=header_row + 1):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    style_data_rows(ws, header_row + 1, header_row + len(df), len(df.columns), df)
    auto_col_width(ws)
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)


def build_summary_sheet(ws, posts: list, df: pd.DataFrame):
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20

    def write_kv(row, key, value, bold_val=False):
        k = ws.cell(row=row, column=1, value=key)
        k.font = Font(bold=True)
        k.fill = PatternFill("solid", fgColor="D9E1F2")
        k.border = CELL_BORDER
        v = ws.cell(row=row, column=2, value=value)
        v.border = CELL_BORDER
        if bold_val:
            v.font = Font(bold=True)

    # Title
    title = ws.cell(row=1, column=1, value="SML Report — Mepolizumab / Nucala")
    title.font = Font(bold=True, size=14, color="1F4E79")
    ws.merge_cells("A1:B1")
    ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    ws.merge_cells("A2:B2")

    row = 4
    write_kv(row, "Total Posts", len(df)); row += 1
    write_kv(row, "Subreddits Covered", df["Subreddit"].nunique()); row += 1
    write_kv(row, "Date Range",
             f"{df['Date'].min()} to {df['Date'].max()}"); row += 1

    row += 1
    ws.cell(row=row, column=1, value="Sentiment Breakdown").font = TITLE_FONT; row += 1
    for sent, count in df["Sentiment"].value_counts().items():
        pct = count / len(df) * 100
        write_kv(row, sent, f"{count} ({pct:.1f}%)"); row += 1

    row += 1
    ws.cell(row=row, column=1, value="Posts by Stakeholder").font = TITLE_FONT; row += 1
    for stake, count in df["Stakeholder"].value_counts().items():
        write_kv(row, stake, count); row += 1

    row += 1
    ws.cell(row=row, column=1, value="Top Emotions").font = TITLE_FONT; row += 1
    for emo, count in df["Emotion"].value_counts().head(5).items():
        write_kv(row, emo, count); row += 1

    row += 1
    ws.cell(row=row, column=1, value="Top Themes").font = TITLE_FONT; row += 1
    all_themes = []
    for t in df["Themes"]:
        all_themes.extend([x.strip() for x in str(t).split(",") if x.strip()])
    for theme, count in Counter(all_themes).most_common(10):
        write_kv(row, theme, count); row += 1

    row += 1
    ws.cell(row=row, column=1, value="Top Drugs Mentioned").font = TITLE_FONT; row += 1
    all_drugs = []
    for d in df["Drugs Mentioned"]:
        all_drugs.extend([x.strip() for x in str(d).split(",") if x.strip()])
    for drug, count in Counter(all_drugs).most_common(10):
        write_kv(row, drug, count); row += 1


def run_export():
    print("Loading analyzed data...")
    posts = load_data()
    df = flatten_posts(posts)

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)

    with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl") as writer:
        # Summary sheet
        df.to_excel(writer, sheet_name="Summary", index=False)  # placeholder
        writer.book.active.title = "_tmp"

        from openpyxl import Workbook
        # We'll build sheets manually via openpyxl after
        pass

    # Rebuild with openpyxl for full styling control
    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    # 1. Summary
    ws_summary = wb.create_sheet("Summary")
    build_summary_sheet(ws_summary, posts, df)

    # 2. All Posts
    ws_all = wb.create_sheet("All Posts")
    write_df_to_sheet(ws_all, df, title="All Posts")

    # 3. Per-stakeholder sheets
    for stakeholder in STAKEHOLDERS:
        filtered = df[df["Stakeholder"] == stakeholder].reset_index(drop=True)
        sheet_name = stakeholder[:31]  # Excel sheet name limit
        ws = wb.create_sheet(sheet_name)
        if filtered.empty:
            ws.cell(row=1, column=1, value=f"No posts found for {stakeholder}")
        else:
            write_df_to_sheet(ws, filtered, title=f"{stakeholder} — {len(filtered)} posts")

    wb.save(OUTPUT_PATH)
    print(f"\nExcel report saved to: {OUTPUT_PATH}")
    print(f"Sheets: Summary, All Posts, {', '.join(STAKEHOLDERS)}")


if __name__ == "__main__":
    run_export()
